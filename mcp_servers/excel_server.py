import json
import os
from pathlib import Path
import openpyxl
from openpyxl.utils import column_index_from_string
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp import types
from dotenv import load_dotenv
import sys

# from mcp_servers.tool_name import ToolNames

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from agents.utils.sandbox import get_safe_path

load_dotenv()
SANDBOX_DIR = Path(os.getenv("AGENT_SANDBOX_DIR", "./agent_workspace")).resolve()

app = Server("excel-mcp-server")

TOOL_REGISTRY = {}
class ToolNames:
    CREATE_WORKBOOK = "create_workbook"
    LIST_SHEETS = "list_sheets"
    READ_SHEET = "read_sheet"
    WRITE_SHEET = "write_sheet"
    SUM_COLUMN = "sum_column"
    AVERAGE_COLUMN = "average_column"
    MIN_MAX_COLUMN = "min_max_column"
    COUNT_COLUMN = "count_column"
    FILTER_ROWS = "filter_rows"
    SORT_SHEET = "sort_sheet"
    FIND_DUPLICATES = "find_duplicates"
    RENAME_SHEET = "rename_sheet"


def tool(name: str, description: str, input_schema: dict):
    def decorator(func):
        TOOL_REGISTRY[name] = {
            "func": func,
            "name": name,
            "description": description,
            "inputSchema": input_schema
        }
        return func

    return decorator

@app.call_tool()
async def call_tool(name: str, arguments: dict):
    handler = TOOL_REGISTRY.get(name)

    if not handler:
        raise ValueError(f"Tool not found: {name}")

    func = handler["func"]
    return func(**arguments)

def sheet_to_json(sheet) -> list[dict]:
    """
    convert a worksheet to a list of dicts using the first row as headers
    return: [
        {"col1": 1, "col2": 2, "col3": 3},
        {"col1": 1, "col2": 2, "col3": 3},
    ]
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    return [
        {headers[i]: row[i] for i in range(len(headers))}
        for row in rows[1:]
    ]


def get_column_values(sheet, column: str) -> list:
    """
    return all non-empty numeric values from a column
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    if column.isalpha() and len(column) <= 3:
        col_idx = column_index_from_string(column.upper()) - 1
    else:
        headers = [str(h).lower() if h else "" for h in rows[0]]
        if column.lower() not in headers:
            raise ValueError(f"Column '{column}' not found. Available: {headers}")
        col_idx = headers.index(column.lower())
        rows = rows[1:]
    return [row[col_idx] for row in rows if row[col_idx] is not None and isinstance(row[col_idx], (int, float))]


def success(data) -> str:
    return json.dumps({"status": "success", "data": data})


def error(msg: str) -> str:
    return json.dumps({"status": "error", "error": msg})


@tool(
    name=ToolNames.CREATE_WORKBOOK,
    description="Create a new Excel workbook with an optional sheet name.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string", "description": "e.g. 'sales.xlsx'"},
            "sheet_name": {"type": "string", "description": "Default: Sheet1"}
        },
        "required": ["filepath"]
    }
)
def create_workbook(filepath: str, sheet_name: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name or "Sheet1"
    wb.save(safe_path)
    result = success(f"Created '{filepath}' with sheet '{ws.title}'")
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.LIST_SHEETS,
    description="List all sheet names in an Excel workbook.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"}
        },
        "required": ["filepath"]
    }
)
def list_sheets(filepath: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    result = success(wb.sheetnames)
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.READ_SHEET,
    description="Read data from an Excel sheet as JSON rows.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "sheet_name": {"type": "string", "description": "Default: first sheet"}
        },
        "required": ["filepath"]
    }
)
def read_sheet(filepath: str, sheet_name: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    result = success(sheet_to_json(ws))
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.WRITE_SHEET,
    description="Write rows of data to an Excel sheet. Overwrites existing content.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "sheet_name": {"type": "string"},
            "rows": {
                "type": "array",
                "description": "List of rows, each row is a list of values. First row = headers.",
                "items": {
                    "type": "array",
                    "items": {
                        "type": ["string", "number"]

                    }
                }
            }
        },
        "required": ["filepath", "rows"]
    }
)
def write_sheet(filepath: str, sheet_name: str, rows: list):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path) if safe_path.exists() else openpyxl.Workbook()
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif sheet_name:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb.active
    ws.delete_rows(1, ws.max_row)
    for row in rows:
        ws.append(row)
    wb.save(safe_path)
    result = success(f"Written {len(rows)} rows to '{filepath}'")
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.SUM_COLUMN,
    description="Calculate the sum of a numeric column.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "sheet_name": {"type": "string"},
            "column": {"type": "string", "description": "Column letter (e.g. 'B') or header name (e.g. 'revenue')"}
        },
        "required": ["filepath", "column"]
    }
)
def sum_column(filepath: str, sheet_name: str, column: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    values = get_column_values(ws, column)
    result = success({"column": column, "sum": sum(values), "row_count": len(values)})
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.AVERAGE_COLUMN,
    description="Calculate the average of a numeric column.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "sheet_name": {"type": "string"},
            "column": {"type": "string"}
        },
        "required": ["filepath", "column"]
    }
)
def average_column(filepath: str, sheet_name: str, column: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    values = get_column_values(ws, column)
    avg = sum(values) / len(values) if values else 0
    result = success({"column": column, "average": round(avg, 4), "row_count": len(values)})
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.MIN_MAX_COLUMN,
    description="Get the minimum and maximum values of a numeric column.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "sheet_name": {"type": "string"},
            "column": {"type": "string"}
        },
        "required": ["filepath", "column"]
    }
)
def min_max_column(filepath: str, sheet_name: str, column: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    values = get_column_values(ws, column)
    result = success({"column": column, "min": min(values), "max": max(values)})
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.COUNT_COLUMN,
    description="Count non-empty values in a column.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "sheet_name": {"type": "string"},
            "column": {"type": "string"}
        },
        "required": ["filepath", "column"]
    }
)
def count_column(filepath: str, sheet_name: str, column: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    values = get_column_values(ws, column)
    result = success({"column": column, "count": len(values)})
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.FILTER_ROWS,
    description="Filter rows where a column matches a condition.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "sheet_name": {"type": "string"},
            "column": {"type": "string"},
            "operator": {
                "type": "string",
                "description": "One of: '>', '<', '>=', '<=', '==', '!=', 'contains'"
            },
            "value": {"description": "Value to compare against"}
        },
        "required": ["filepath", "column", "operator", "value"]
    }
)
def filter_rows(filepath: str, sheet_name: str, column: str, operation: str, value: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = sheet_to_json(ws)
    col = column
    op = operation
    val = value

    def matches(row):
        cell = row.get(col)
        if cell is None:
            return False
        try:
            if op == ">":
                return float(cell) > float(val)
            elif op == "<":
                return float(cell) < float(val)
            elif op == ">=":
                return float(cell) >= float(val)
            elif op == "<=":
                return float(cell) <= float(val)
            elif op == "==":
                return str(cell) == str(val)
            elif op == "!=":
                return str(cell) != str(val)
            elif op == "contains":
                return str(val).lower() in str(cell).lower()
        except:
            return False
        return False

    filtered = [r for r in rows if matches(r)]
    result = success({"matched_rows": len(filtered), "rows": filtered})
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.SORT_SHEET,
    description="Sort sheet rows by a column.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "sheet_name": {"type": "string"},
            "column": {"type": "string"},
            "ascending": {"type": "boolean", "description": "Default: true"}
        },
        "required": ["filepath", "column"]
    }
)
def sort_sheet(filepath: str, sheet_name: str, column: str, ascending: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = sheet_to_json(ws)
    col = column
    ascending = ascending
    sorted_rows = sorted(rows, key=lambda r: (r.get(col) is None, r.get(col)), reverse=not ascending)

    # Write sorted data back
    ws.delete_rows(1, ws.max_row)
    if sorted_rows:
        ws.append(list(sorted_rows[0].keys()))  # headers
        for row in sorted_rows:
            ws.append(list(row.values()))
    wb.save(safe_path)
    result = success(f"Sorted by '{col}' {'ascending' if ascending else 'descending'}")
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.FIND_DUPLICATES,
    description="Find duplicate values in a column.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "sheet_name": {"type": "string"},
            "column": {"type": "string"}
        },
        "required": ["filepath", "column"]
    }
)
def find_duplicates(filepath: str, sheet_name: str, column: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    values = [str(v) for v in get_column_values(ws, column)]
    seen = set()
    duplicates = set()
    for v in values:
        if v in seen:
            duplicates.add(v)
        seen.add(v)
    result = success({"duplicate_values": list(duplicates), "count": len(duplicates)})
    return [types.TextContent(type="text", text=result)]


@tool(
    name=ToolNames.RENAME_SHEET,
    description="Rename a sheet in a workbook.",
    input_schema={
        "type": "object",
        "properties": {
            "filepath": {"type": "string"},
            "old_name": {"type": "string"},
            "new_name": {"type": "string"}
        },
        "required": ["filepath", "old_name", "new_name"]
    }
)
def rename_sheet(filepath: str, old_name: str, new_name: str):
    safe_path = get_safe_path(filepath)
    wb = openpyxl.load_workbook(safe_path)
    ws = wb[old_name]
    ws.title = new_name
    wb.save(safe_path)
    result = success(f"Renamed '{old_name}' to '{new_name}'")
    return [types.TextContent(type="text", text=result)]


@app.list_tools()
async def list_tools() -> list[types.Tool]:
    tools = []
    for name, meta in TOOL_REGISTRY.items():
        tools.append(
            types.Tool(
                name=name,
                description=meta["description"],
                inputSchema=meta["inputSchema"]
            )
        )
    return tools


async def main():
    async with stdio_server() as (read_stream, write_stream):
        await app.run(read_stream, write_stream, app.create_initialization_options())


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())